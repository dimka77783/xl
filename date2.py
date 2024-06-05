import openpyxl

path = "12.xlsx"
wb_obj = openpyxl.load_workbook(path)  # Открываем файл
sheet_obj = wb_obj["работа оборудования"]
sheet_obj1 = wb_obj["мощность"]# Выбираем активный лист таблицы(
m_row = sheet_obj.max_row
m_row1 = sheet_obj1.max_row
#print(m_row)

date_start_list = []
date_stop_list = []

for i in range(2, m_row + 1):  # для всех значений из листа для свода
    cell_ob = sheet_obj.cell(row=i, column=4)  # дата начала из листа свод2
    data_start = cell_ob.value
    cell_obj1 = sheet_obj.cell(row=i, column=5)  # дата начала из листа свод2
    data_stop = cell_obj1.value
    cell_obj2 = sheet_obj.cell(row=i, column=2)
    year = cell_obj2.value
    oo = data_start.split(' ')
    aa = data_stop.split(' ')
    date = str(oo[0])+'.'+str(year) + ' ' + str(oo[1])+':00'
    date_start_list.append(str(date))
    date2 = str(aa[0]) + '.' + str(year) + ' ' + str(aa[1]) + ':00'
    date_stop_list.append(str(date2))

for x, statN in enumerate(date_start_list):
    sheet_obj.cell(row=x+2, column=15).value = statN
for x, statN in enumerate(date_stop_list):
    sheet_obj.cell(row=x+2, column=16).value = statN

for i in range(2, m_row + 1):  # для всех значений из листа для свода
    k = sheet_obj1.cell(row=i, column=7)  # дата начала из листа свод2
    k.value = '=TEXT(A2,"dd.mm.yyyy HH:MM:SS")'

wb_obj.save("12.xlsx")


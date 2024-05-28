import openpyxl
from statistics import mean
from openpyxl.workbook.defined_name import DefinedName

path = "123456789.xlsx"
wb_obj = openpyxl.load_workbook(path)  # Открываем файл
sheet_obj1 = wb_obj["для свода 2"]
sheet_obj2 = wb_obj["мощность"]  # Выбираем активный лист таблицы(
m_row1 = sheet_obj1.max_row
m_row2 = sheet_obj2.max_row
#print(m_row1)
date_start_list = []
date_stop_list = []
date_power_list = []
power_now_list = []
index_stop_list = []
index_start_list = []
delta_power_list = []
mean_power_list = []
count = 1
for i in range(2, m_row1 + 1):  # для всех значений из листа для свода
    cell_ob = sheet_obj1.cell(row=i, column=2)  # дата начала из листа свод2
    data_start = cell_ob.value
    # $print(data_start)
    date_start_list.append(str(data_start))
    cell_ob2 = sheet_obj1.cell(row=i, column=3)  # дата окончания из листа свод2
    data_stop = cell_ob2.value
    date_stop_list.append(str(data_stop))  # формируем список дат окончания
# print(date_start_list)

for n in range(2, m_row2 + 1):
    cell_obj2 = sheet_obj2.cell(row=n, column=1)  # мощность момент из листа мощность
    date_power = cell_obj2.value
    date_power_list.append(str(date_power))
    # cell_obj3 = sheet_obj2.cell(row=n, column=2)  # мощность момент из листа мощность
    # power_now = cell_obj3.value
    # power_now_list.append(str(power_now)) # формируем список мощностей текущих
# print(date_power_list)

for date in date_start_list:
    index_start = date_power_list.index(date) + 2  # получаем номер строки даты начала
    index_start_list.append(str(index_start))

# print(index_start_list)
for date in date_stop_list:
    index_stop = date_power_list.index(date) + 2  # получаем номер строки даты конца
    index_stop_list.append(str(index_stop))
# print(int(index_stop_list))
for num in index_start_list:
    file = open("otus.txt", "w")
    file.write(str(num))
    file.close()


for start, stop in zip(index_start_list, index_stop_list): # получаем список всех дельт
    for c in range(int(start), int(stop) + 1):
        cell_obj7 = sheet_obj2.cell(row=c, column=4)
        delta_power = cell_obj7.value
        delta_power_list.append(float(delta_power))         # получаем список дельт
        mean = sum(delta_power_list)/len(delta_power_list)  # получаем ср знач из диапазона дельт

    mean_power_list.append(mean)                      # собираем средние значения в список
    delta_power_list.clear()                          # очишаем список для новых пар значений

for i, statN in enumerate(mean_power_list):
    sheet_obj1.cell(row=i+2, column=15).value = statN  # вносим средние значения в таблицу


wb_obj.save("123456789.xlsx")

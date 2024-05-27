import openpyxl

path = "Работа с низкой мощностью(3).xlsx"
wb_obj = openpyxl.load_workbook(path)  # Открываем файл
sheet_obj1 = wb_obj["для свода 2"]
sheet_obj2 = wb_obj["мощность"]       # Выбираем активный лист таблицы(
m_row1 = sheet_obj1.max_row
m_row2 = sheet_obj2.max_row
date_start_list = []
date_stop_list = []
date_power_list = []
power_now_list = []
index_stop_list = []
index_start_list = []
count = 1
for i in range(2, m_row1 + 1):  # для всех значений из листа для свода
    cell_ob = sheet_obj1.cell(row=i, column=2)  # дата начала из листа свод2
    data_start = cell_ob.value
    date_start_list.append(str(data_start))
    cell_ob2 = sheet_obj1.cell(row=i, column=3)  # дата окончания из листа свод2
    data_stop = cell_ob2.value
    date_stop_list.append(str(data_stop)) # формируем список дат окончания

for n in range(2, m_row2 + 1):
    cell_obj2 = sheet_obj2.cell(row=n, column=1)  # мощность момент из листа мощность
    date_power = cell_obj2.value
    date_power_list.append(str(date_power))
    cell_obj3 = sheet_obj2.cell(row=n, column=2)  # мощность момент из листа мощность
    power_now = cell_obj3.value
    power_now_list.append(str(power_now)) # формируем список мощностей текущих
print(date_start_list)
for date in date_start_list:
    index_start = date_power_list.index(date)+2 # получаем номер строки
    index_start_list.append(str(index_start))
    print(index_start)
for date in date_stop_list:
    index_stop = date_power_list.index(date)+2 # получаем номер строки
    index_stop_list.append(str(index_stop))


#проверка22

#print(date_list)
#cell_obj = sheet_obj2.cell(row=i, column=1)  # дата из листа мощность
#date2 = cell_obj.value
for n in range(3, m_row2 + 1):
    cell_obj2 = sheet_obj2.cell(row=n, column=2)  # мощность момент из листа мощность
    power_nom = cell_obj2.value
    count = count + 1
    if isinstance(power_nom, float): # если мощность число с плавающей точкой
        pass# для всех значений из листа мощность

        #print(str(count)+" "+str(power_nom))
        #if power_nom is float:
        #print(str(count)+" "+str(power_nom))
    else:
        pass# print(str(count) + "не число")

#sheats = wb_obj.sheetnames
#for sheet in sheats:
    #print(sheet)


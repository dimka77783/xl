import openpyxl
path = "1234.xlsx"
wb_obj = openpyxl.load_workbook(path)  # Открываем файл
sheet_obj1 = wb_obj["мощность вся 1-5"]
sheet_obj2 = wb_obj["работа оборудования"]  # Выбираем активный лист таблицы(
m_row1 = sheet_obj1.max_row
m_row2 = sheet_obj2.max_row

date_start_list = []
date_stop_list = []
date_power_list = []
power_now_list = []
index_stop_list = []
index_start_list = []
delta_power_list = []
mean_power_list = []
"""
получаем список date_start_list и date_stop_list
из листа работа оборудования колонки 4 и 5
"""
for i in range(2, m_row2 + 1):  # для всех значений из листа для свода
    cell_ob = sheet_obj2.cell(row=i, column=4)  # дата начала из листа свод2
    data_start = cell_ob.value
    date_start_list.append(str(data_start))
    cell_ob2 = sheet_obj2.cell(row=i, column=5)  # дата окончания из листа свод2
    data_stop = cell_ob2.value
    date_stop_list.append(str(data_stop))

"""
получаем список date_power_list из листа мощность вся 1-5 
колонка 1
"""

for n in range(2, m_row1 + 1):
    cell_obj2 = sheet_obj1.cell(row=n, column=1)  # мощность момент из листа мощность
    date_power = cell_obj2.value
    date_power_list.append(str(date_power))

"""
получаем список номеров строк из списка date_stop_list и date_start_list
"""
for date in date_start_list:
    if date in date_power_list:
        index_start = date_power_list.index(date) + 2  # получаем номер строки даты начала
        index_start_list.append(str(index_start))
    else:
        print(date+"начало")

for date in date_stop_list:
    if date in date_power_list:
        index_stop = date_power_list.index(date) + 2  # получаем номер строки даты конца
        index_stop_list.append(str(index_stop))
    else:
        print(date+"конец")

print(len(index_start_list))
print(len(index_stop_list))
"""
закрашивае ячейки в заданных диапазонах start, stop
"""
for start, stop in zip(index_start_list, index_stop_list):
    for n in range(int(start), int(stop) + 1):# закрашиваем диапазоны
        cell_obj7 = sheet_obj1.cell(row=n, column=1)
        cell_obj7.fill = openpyxl.styles.PatternFill(start_color='0070c1', end_color='0070c1', fill_type='solid')


wb_obj.save("1234.xlsx")